#!/usr/bin/env python3
"""
CostReports360 - Local CLI Tool
Generates AWS cost comparison reports locally using AWS CLI credentials.
"""

import argparse
import boto3
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Cost change thresholds (as percentages)
MINIMAL_CHANGE_THRESHOLD = 5  # Changes below this percentage are considered minimal
HIGH_CHANGE_THRESHOLD = 20    # Changes above this are highlighted as significant
MIN_SIGNIFICANT_COST = 0.01   # Minimum cost difference to be considered significant


def get_aws_session(profile=None, region='us-east-1'):
    """Create AWS session using local credentials."""
    try:
        if profile:
            session = boto3.Session(profile_name=profile, region_name=region)
        else:
            session = boto3.Session(region_name=region)
        
        # Verify credentials and display authentication info
        sts = session.client('sts')
        identity = sts.get_caller_identity()
        print(f"✓ Authenticated as: {identity['Arn']}")
        print(f"✓ Account ID: {identity['Account']}")
        return session
    except Exception as e:
        print(f"✗ Authentication failed: {e}")
        print("\nMake sure you have configured AWS CLI credentials:")
        print("  aws configure")
        print("\nOr specify a profile:")
        print("  python cost_report_cli.py --profile <profile_name> ...")
        sys.exit(1)


def validate_months(months):
    """Validate month format and count."""
    if len(months) < 2:
        print("✗ Error: At least 2 months are required")
        sys.exit(1)
    if len(months) > 6:
        print("✗ Error: Maximum 6 months allowed")
        sys.exit(1)
    
    validated = []
    for month in months:
        try:
            datetime.strptime(month, "%Y-%m")
            validated.append(month)
        except ValueError:
            print(f"✗ Error: Invalid month format '{month}'. Use YYYY-MM (e.g., 2024-01)")
            sys.exit(1)
    
    return sorted(validated)


def fetch_cost_data(session, months):
    """Fetch cost data from AWS Cost Explorer."""
    ce = session.client('ce')
    
    service_costs = {}
    regional_costs = {}
    
    for month in months:
        print(f"  Fetching data for {month}...")
        start = f"{month}-01"
        end_date = datetime.strptime(start, "%Y-%m-%d")
        if end_date.month == 12:
            end = f"{end_date.year + 1}-01-01"
        else:
            end = f"{end_date.year}-{str(end_date.month + 1).zfill(2)}-01"
        
        # Get detailed usage data
        response = ce.get_cost_and_usage(
            TimePeriod={'Start': start, 'End': end},
            Granularity='MONTHLY',
            Metrics=['NetUnblendedCost', 'UsageQuantity'],
            GroupBy=[
                {'Type': 'DIMENSION', 'Key': 'SERVICE'},
                {'Type': 'DIMENSION', 'Key': 'USAGE_TYPE'}
            ],
            Filter={'Not': {'Dimensions': {'Key': 'RECORD_TYPE', 'Values': ['Tax']}}}
        )
        
        for result in response['ResultsByTime'][0]['Groups']:
            service = result['Keys'][0]
            usage_type = result['Keys'][1]
            cost = float(result['Metrics']['NetUnblendedCost']['Amount'])
            usage = float(result['Metrics']['UsageQuantity']['Amount'])
            
            if cost > 0:
                if service not in service_costs:
                    service_costs[service] = {}
                if month not in service_costs[service]:
                    service_costs[service][month] = {'total': 0, 'details': []}
                
                service_costs[service][month]['total'] += cost
                service_costs[service][month]['details'].append({
                    'usage_type': usage_type,
                    'cost': cost,
                    'usage': usage
                })
        
        # Get regional costs
        regional_response = ce.get_cost_and_usage(
            TimePeriod={'Start': start, 'End': end},
            Granularity='MONTHLY',
            Metrics=['NetUnblendedCost'],
            GroupBy=[{'Type': 'DIMENSION', 'Key': 'REGION'}],
            Filter={'Not': {'Dimensions': {'Key': 'RECORD_TYPE', 'Values': ['Tax']}}}
        )
        
        for result in regional_response['ResultsByTime'][0]['Groups']:
            region = result['Keys'][0]
            cost = float(result['Metrics']['NetUnblendedCost']['Amount'])
            
            if cost > 0:
                if region not in regional_costs:
                    regional_costs[region] = {}
                regional_costs[region][month] = cost
    
    return service_costs, regional_costs


def categorize_services(sorted_services, months):
    """Categorize services by cost change."""
    increased_services = []
    decreased_services = []
    same_services = []
    
    for service, data in sorted_services:
        month_costs = [data.get(m, {}).get('total', 0) for m in months]
        if len(month_costs) >= 2:
            change = month_costs[-1] - month_costs[0]
            pct = (change / month_costs[0] * 100) if month_costs[0] > 0 else 0
            
            if abs(pct) < MINIMAL_CHANGE_THRESHOLD:
                same_services.append((service, data))
            elif change > 0:
                increased_services.append((service, data))
            else:
                decreased_services.append((service, data))
        else:
            same_services.append((service, data))
    
    return increased_services, decreased_services, same_services


def format_compute_usage(usage_type, cost, usage):
    """Format compute instance usage with hourly rates."""
    compute_patterns = [
        ('BoxUsage:', 'EC2'),
        ('HeavyUsage:', 'EC2 Reserved'),
        ('SpotUsage:', 'EC2 Spot'),
        ('InstanceUsage:', 'RDS'),
        ('Multi-AZUsage:', 'RDS Multi-AZ'),
        ('ServerlessUsage:', 'RDS Serverless'),
        ('NodeUsage:', 'ElastiCache'),
        ('Node:', 'Redshift'),
    ]
    
    for pattern, service_type in compute_patterns:
        if pattern in usage_type:
            instance_type = usage_type.split(pattern)[1].split(':')[0]
            if usage > 0:
                hourly_rate = cost / usage
                return f"{instance_type} ({usage:,.3f} Hrs @ ${hourly_rate:.4f}): ${cost:,.2f}"
            else:
                return f"{instance_type}: ${cost:,.2f}"
    
    return None


def generate_detailed_comparison(month_names, data, months):
    """Generate detailed comparison text for a service."""
    lines = []
    
    for i, month in enumerate(months):
        month_data = data.get(month, {})
        
        if 'details' in month_data:
            lines.append(f"\n[{month_names[i].upper()} BREAKDOWN]")
            
            sorted_details = sorted(month_data['details'], key=lambda x: x['cost'], reverse=True)[:5]
            for detail in sorted_details:
                formatted = format_compute_usage(detail['usage_type'], detail['cost'], detail['usage'])
                
                if formatted:
                    lines.append(formatted)
                else:
                    lines.append(f"{detail['usage_type']}: USD {detail['cost']:,.2f}")
                    if detail['usage'] > 0:
                        lines.append(f"Usage: {detail['usage']:,.3f} units")
    
    if len(months) >= 2:
        first_total = data.get(months[0], {}).get('total', 0)
        last_total = data.get(months[-1], {}).get('total', 0)
        change = last_total - first_total
        lines.append(f"\n[COST DIFFERENCE]")
        lines.append(f"USD {abs(change):,.2f} ({'Increased' if change > 0 else 'Decreased'})")
    
    return '\n'.join(lines)


def generate_total_comparison(month_names, totals):
    """Generate total comparison text."""
    lines = []
    for i, name in enumerate(month_names):
        lines.append(f"{name} Total: USD {totals[i]:,.2f}")
    
    if len(totals) >= 2:
        change = totals[-1] - totals[0]
        lines.append(f"\nTotal Change: USD {abs(change):,.2f} ({'Increased' if change > 0 else 'Decreased'})")
    
    return '\n'.join(lines)


def generate_detailed_reason(month_names, data, month_costs, months):
    """Generate detailed reason for cost changes."""
    if len(month_costs) < 2:
        return "Insufficient data"
    
    change = month_costs[-1] - month_costs[0]
    pct = (change / month_costs[0] * 100) if month_costs[0] > 0 else 0
    
    if abs(pct) < MINIMAL_CHANGE_THRESHOLD:
        return "Minimal Cost Difference"
    
    first_month = months[0]
    last_month = months[-1]
    first_details = {d['usage_type']: d['cost'] for d in data.get(first_month, {}).get('details', [])}
    last_details = {d['usage_type']: d['cost'] for d in data.get(last_month, {}).get('details', [])}
    
    changes = []
    all_types = set(first_details.keys()) | set(last_details.keys())
    for usage_type in all_types:
        first_cost = first_details.get(usage_type, 0)
        last_cost = last_details.get(usage_type, 0)
        if first_cost > 0 or last_cost > 0:
            diff = last_cost - first_cost
            changes.append((usage_type, diff, first_cost, last_cost))
    
    changes.sort(key=lambda x: abs(x[1]), reverse=True)
    
    lines = [f"Cost {'increased' if change > 0 else 'decreased'} by USD {abs(change):,.2f} ({abs(pct):.1f}%)"]
    
    significant_changes = [c for c in changes[:3] if abs(c[1]) > MIN_SIGNIFICANT_COST]
    if significant_changes:
        lines.append("\nTop changes:")
        for usage_type, diff, first, last in significant_changes:
            lines.append(f"- {usage_type}: USD {first:,.2f} → USD {last:,.2f}")
    
    return '\n'.join(lines)


def generate_simple_reason(costs):
    """Generate simple reason for cost changes."""
    if len(costs) < 2:
        return "Insufficient data"
    
    change = costs[-1] - costs[0]
    pct = (change / costs[0] * 100) if costs[0] > 0 else 0
    
    if abs(pct) < MINIMAL_CHANGE_THRESHOLD:
        return "Minimal Cost Difference"
    elif change > 0:
        return f"Cost increased by USD {abs(change):,.2f} ({abs(pct):.1f}% increase)"
    else:
        return f"Cost decreased by USD {abs(change):,.2f} ({abs(pct):.1f}% decrease)"


def create_service_sheet(ws, sorted_services, months, month_names):
    """Create a service costs sheet."""
    # Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    light_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    light_red = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    dark_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header
    headers = ['Services'] + month_names + ['Service Total', 'Comparison', 'Reason']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 40
    
    # Data rows
    row = 2
    month_totals = [0.0] * len(months)
    
    for service, data in sorted_services:
        cell = ws.cell(row, 1, service)
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        month_costs = []
        for col, month in enumerate(months, 2):
            cost = data.get(month, {}).get('total', 0)
            month_costs.append(cost)
            month_totals[col - 2] += cost
            cell = ws.cell(row, col, round(cost, 2))
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        total = sum(month_costs)
        cell = ws.cell(row, len(months) + 2, round(total, 2))
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        comparison = generate_detailed_comparison(month_names, data, months)
        cell = ws.cell(row, len(months) + 3, comparison)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = Font(size=10)
        
        reason = generate_detailed_reason(month_names, data, month_costs, months)
        cell = ws.cell(row, len(months) + 4, reason)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Color coding
        if len(month_costs) >= 2:
            change = month_costs[-1] - month_costs[0]
            pct = (change / month_costs[0] * 100) if month_costs[0] > 0 else 0
            
            if abs(pct) < MINIMAL_CHANGE_THRESHOLD:
                fill = blue_fill
            elif change > 0:
                fill = dark_red if pct > HIGH_CHANGE_THRESHOLD else light_red
            else:
                fill = light_green
            
            for col in range(1, len(months) + 5):
                ws.cell(row, col).fill = fill
        
        num_lines = comparison.count('\n') + 1
        ws.row_dimensions[row].height = (num_lines * 15) + 20
        
        row += 1
    
    # Total row
    cell = ws.cell(row, 1, "Total")
    cell.font = Font(bold=True)
    cell.border = border
    cell.fill = yellow_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    grand_total = 0.0
    for col, total in enumerate(month_totals, 2):
        cell = ws.cell(row, col, round(total, 2))
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        grand_total += total
    
    cell = ws.cell(row, len(months) + 2, round(grand_total, 2))
    cell.font = Font(bold=True)
    cell.border = border
    cell.fill = yellow_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    total_comparison = generate_total_comparison(month_names, month_totals)
    cell = ws.cell(row, len(months) + 3, total_comparison)
    cell.border = border
    cell.font = Font(bold=True)
    cell.fill = yellow_fill
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    total_reason = generate_simple_reason(month_totals)
    cell = ws.cell(row, len(months) + 4, total_reason)
    cell.border = border
    cell.font = Font(bold=True)
    cell.fill = yellow_fill
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    num_lines = total_comparison.count('\n') + 1
    ws.row_dimensions[row].height = (num_lines * 15) + 20
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    for col in range(2, len(months) + 3):
        ws.column_dimensions[chr(64 + col)].width = 20
    ws.column_dimensions[chr(64 + len(months) + 3)].width = 50
    ws.column_dimensions[chr(64 + len(months) + 4)].width = 65


def create_regional_sheet(ws, regional_costs, months, month_names):
    """Create regional costs sheet."""
    # Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header
    headers = ['Region'] + month_names + ['Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 40
    
    sorted_regions = sorted(
        regional_costs.items(),
        key=lambda x: sum(x[1].values()),
        reverse=True
    )
    
    row = 2
    month_totals = [0.0] * len(months)
    
    for region, costs in sorted_regions:
        region_total = sum(costs.get(month, 0) for month in months)
        
        cell = ws.cell(row, 1, region)
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = green_fill
        
        for col, month in enumerate(months, 2):
            cost = costs.get(month, 0)
            month_totals[col - 2] += cost
            cell = ws.cell(row, col, round(cost, 2))
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = green_fill
        
        cell = ws.cell(row, len(months) + 2, round(region_total, 2))
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = green_fill
        
        row += 1
    
    # Total row
    cell = ws.cell(row, 1, "Total")
    cell.font = Font(bold=True)
    cell.border = border
    cell.fill = yellow_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    grand_total = 0.0
    for col, total in enumerate(month_totals, 2):
        cell = ws.cell(row, col, round(total, 2))
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        grand_total += total
    
    cell = ws.cell(row, len(months) + 2, round(grand_total, 2))
    cell.font = Font(bold=True)
    cell.border = border
    cell.fill = yellow_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, len(months) + 3):
        ws.column_dimensions[chr(64 + col)].width = 15


def generate_report(service_costs, regional_costs, months, client_name, output_dir):
    """Generate the Excel report."""
    # Convert to month names
    month_names = []
    month_names_short = []
    for month in months:
        dt = datetime.strptime(month, "%Y-%m")
        month_names.append(dt.strftime("%B %Y"))
        month_names_short.append(dt.strftime("%B"))
    
    # Sort services by total cost
    sorted_services = sorted(
        service_costs.items(),
        key=lambda x: sum(x[1].get(m, {}).get('total', 0) for m in months),
        reverse=True
    )
    
    # Categorize services
    increased, decreased, same = categorize_services(sorted_services, months)
    
    # Generate Excel
    wb = Workbook()
    
    # Sheet 1: Complete Service Costs
    ws1 = wb.active
    ws1.title = "Complete Service Costs"
    create_service_sheet(ws1, sorted_services, months, month_names)
    
    # Sheet 2: Increased Service Costs
    ws2 = wb.create_sheet("Increased Service Costs")
    create_service_sheet(ws2, increased, months, month_names)
    
    # Sheet 3: Decreased Service Costs
    ws3 = wb.create_sheet("Decreased Service Costs")
    create_service_sheet(ws3, decreased, months, month_names)
    
    # Sheet 4: Same Service Costs
    ws4 = wb.create_sheet("Same Service Costs")
    create_service_sheet(ws4, same, months, month_names)
    
    # Sheet 5: Regional Costs
    ws5 = wb.create_sheet("Per-region Costs")
    create_regional_sheet(ws5, regional_costs, months, month_names)
    
    # Generate filename
    client_name_formatted = client_name.replace(' ', '-')
    months_str = '-'.join(month_names_short)
    filename = f"{client_name_formatted}-{months_str}-CostReport.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Save
    wb.save(filepath)
    
    return filepath


def main():
    parser = argparse.ArgumentParser(
        description='CostReports360 - Generate AWS cost comparison reports locally',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate report for the last 3 months
  python cost_report_cli.py --client "Acme Corp" --months 2024-10 2024-11 2024-12

  # Use a specific AWS profile
  python cost_report_cli.py --profile production --client "Client A" --months 2024-09 2024-10

  # Save report to specific directory
  python cost_report_cli.py --client "Client B" --months 2024-11 2024-12 --output /path/to/reports
        """
    )
    
    parser.add_argument(
        '--client', '-c',
        required=True,
        help='Client name (used in report filename)'
    )
    
    parser.add_argument(
        '--months', '-m',
        nargs='+',
        required=True,
        help='Months to compare (2-6 months, format: YYYY-MM)'
    )
    
    parser.add_argument(
        '--profile', '-p',
        default=None,
        help='AWS CLI profile to use (default: default profile)'
    )
    
    parser.add_argument(
        '--region', '-r',
        default='us-east-1',
        help='AWS region (default: us-east-1, Cost Explorer API is global but us-east-1 is recommended)'
    )
    
    parser.add_argument(
        '--output', '-o',
        default='.',
        help='Output directory for the report (default: current directory)'
    )
    
    args = parser.parse_args()
    
    print("\n" + "="*60)
    print("CostReports360 - Local Cost Report Generator")
    print("="*60 + "\n")
    
    # Validate months
    print("▶ Validating parameters...")
    months = validate_months(args.months)
    print(f"✓ Months to analyze: {', '.join(months)}")
    print(f"✓ Client name: {args.client}")
    
    # Check output directory
    if not os.path.exists(args.output):
        os.makedirs(args.output)
        print(f"✓ Created output directory: {args.output}")
    else:
        print(f"✓ Output directory: {args.output}")
    
    # Authenticate
    print("\n▶ Authenticating with AWS...")
    session = get_aws_session(profile=args.profile, region=args.region)
    
    # Fetch data
    print("\n▶ Fetching cost data from AWS Cost Explorer...")
    try:
        service_costs, regional_costs = fetch_cost_data(session, months)
    except Exception as e:
        print(f"\n✗ Error fetching cost data: {e}")
        print("\nPossible causes:")
        print("  - Cost Explorer is not enabled in your AWS account")
        print("  - Insufficient IAM permissions (need ce:GetCostAndUsage)")
        print("  - Invalid AWS credentials")
        sys.exit(1)
    
    print(f"✓ Found {len(service_costs)} services with costs")
    print(f"✓ Found {len(regional_costs)} regions with costs")
    
    # Generate report
    print("\n▶ Generating Excel report...")
    try:
        filepath = generate_report(service_costs, regional_costs, months, args.client, args.output)
        print(f"✓ Report saved to: {filepath}")
    except Exception as e:
        print(f"\n✗ Error generating report: {e}")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("✓ Report generation complete!")
    print("="*60 + "\n")


if __name__ == '__main__':
    main()
